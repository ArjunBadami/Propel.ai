import openpyxl.styles
import pandas as pd
import os
from datetime import datetime, timedelta
import openpyxl
import re

def read_excel_tasks(filename):
    # Load the Excel file
    #filename = filename + ".xlsx"
    df = pd.read_excel(filename, sheet_name= "Sheet1", engine='openpyxl')

    # Initialize an empty list to store task data
    tasks = []

    # Iterate over each row in the DataFrame
    for index, row in df.iterrows():
        # Assuming the structure: Task ID | Task Name | Task Duration | Dependencies
        task_id = int(row.iloc[0])  # Convert Task ID to integer
        task_name = str(row.iloc[1])  # Task Name as string
        task_duration = int(row.iloc[4])  # Task Duration as integer
        dependencies = [int(v) for v in str(row.iloc[3]).split(',')] if (pd.notna(row.iloc[3])
                                                                         and str(row.iloc[3]).strip() != '') else []
        #print(str(row.iloc[2]))
        start_time = ''
        if(pd.notna(row.iloc[2]) and str(row.iloc[2]).strip() != ''):
            start_time = datetime.strptime(str(row.iloc[2]), "%Y-%m-%d %H:%M:%S")
        # Append the task as a list to the tasks list
        tasks.append([task_id, task_name, task_duration, dependencies, start_time])

    return tasks


def convert_tasks_for_gantt(tasks):
    newtasks = []
    i = 0
    for t in tasks:
        task = {}
        task["Task"] = t[0]
        task["Name"] = t[1]
        task["Start"] = t[4]
        task["Duration"] = t[2]
        task["Dependencies"] = t[3]
        task["Finish"] = t[4] + timedelta(days=int(t[2]))
        newtasks.append(task)
        i += 1
        if i >= 25:
            break
    
    return newtasks


def getprojectdetails(project_path):
    filename = project_path
    # Load the Excel file
    df = pd.read_excel(filename, sheet_name='Details',engine='openpyxl', header=None)
    details = {}
    for index, row in df.iterrows():
        key = str(row.iloc[0])
        value = str(row.iloc[1])
        details[key] = value

    return details


def getprojectsections(project_path, sheetname):
    filename = project_path
    # Load the Excel file
    df = pd.read_excel(filename, sheet_name=sheetname,engine='openpyxl', header=None)
    items = []
    for index, row in df.iterrows():
        items.append(str(row.iloc[0]))

    return items


def write_critical_start_times(project_path, tasks):
    if(os.path.exists(project_path) == False):
        return "File not found"
    wb = openpyxl.load_workbook(project_path)
    sheet = wb['Sheet1']
    df = pd.read_excel(project_path, sheet_name= "Sheet1", engine='openpyxl')
    light_yellow_fill = openpyxl.styles.PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")
    # Iterate over each row in the DataFrame
    for index, row in df.iterrows():
        # Assuming the structure: Task ID | Task Name | Task Duration | Dependencies
        task_id = int(row.iloc[0])  # Convert Task ID to integer
        start_time = ''
        if(pd.notna(row.iloc[2]) == False or str(row.iloc[2]).strip() == ''):
            start_time = tasks['task' + str(task_id)]['ES']
            #start_time = start_time.strftime("%m/%d/%Y")
            #start_time = start_time.strftime("%Y-%m-%d %H:%M:%S")
            #sheet.cell(row=(index+2), column=3, value=str(start_time))
            sheet['C' + str(index+2)] = start_time
            #date_stylez = openpyxl.styles.NamedStyle(name='date_style', number_format='M/D/YYYY')
            sheet['C' + str(index+2)].number_format = 'M/D/YYYY'
            sheet['C' + str(index+2)].fill = light_yellow_fill
            sheet['C' + str(index+2)].alignment = openpyxl.styles.Alignment(horizontal='left')


    wb.save(project_path)
    return "Successfully generated critical path for: " + project_path


def check_and_extract_invoice_number(text):
    # Define the pattern
    pattern = r"^Invoice (\d+) update \(based on contract\)$"
    
    # Use re.match to check if the string matches the pattern
    match = re.match(pattern, text)
    
    if match:
        # Extract the integer using group(1)
        invoice_number = int(match.group(1))
        return invoice_number
    else:
        return None
    

def getmilestone_dates(project_path):
    dates = {}
    if(os.path.exists(project_path) == False):
        return dates
    wb = openpyxl.load_workbook(project_path)
    sheet_name = 'Sheet1'
    if sheet_name not in wb.sheetnames:
        wb.create_sheet(sheet_name)
    df = pd.read_excel(project_path, sheet_name=sheet_name,engine='openpyxl')
    for index, row in df.iterrows():
        task_name = str(row.iloc[1])
        start_time = ''
        if(pd.notna(row.iloc[2]) and str(row.iloc[2]).strip() != ''):
            start_time = datetime.strptime(str(row.iloc[2]), "%Y-%m-%d %H:%M:%S")
        
        num = check_and_extract_invoice_number(task_name)
        if num is not None and start_time != '':
            dates[num] = start_time
        
    return dates


def getmilestone_dates_pss():
    dates = {}
    project_path = r'.\Sample_Project\rwservlet.xlsx'
    if(os.path.exists(project_path) == False):
        return dates
    wb = openpyxl.load_workbook(project_path)
    sheet_name = 'Sheet1'
    if sheet_name not in wb.sheetnames:
        wb.create_sheet(sheet_name)
    df = pd.read_excel(project_path, sheet_name=sheet_name,engine='openpyxl')
    filtered_row = df[df['PD_NO'] == "2-4F14"]
    if not filtered_row.empty:
        single_row = filtered_row.iloc[0]
        key = "BASLAY_PLAN"
        if key in single_row:
            dates[1] = single_row[key]
            if pd.api.types.is_datetime64_any_dtype(dates[1]):
                dates[1] = dates[1].strftime('%Y-%m-%d')
        key = 'PKG_TEST_ACTUAL'
        if key in single_row:
            dates[7] = single_row[key]
            if pd.api.types.is_datetime64_any_dtype(dates[7]):
                dates[7] = dates[7].strftime('%Y-%m-%d')
    return dates
        

def write_payment_milestones(project_path, po_data):
    if(os.path.exists(project_path) == False):
        return "File not found"
    wb = openpyxl.load_workbook(project_path)
    sheet_name = 'Payments'
    if sheet_name not in wb.sheetnames:
        wb.create_sheet(sheet_name)
    sheet = wb[sheet_name]
    points = po_data[0]
    if points.endswith('TERMINATE'):
        points = points[:-len('TERMINATE')]
    points = points.split('##')
    sheet['A1'] = 'No.'
    sheet['B1'] = 'MILESTONE DESCRIPTION'
    sheet['C1'] = '%'
    sheet['D1'] = 'STATUS'
    sheet['E1'] = 'Forecast Date'
    i = 1
    dates = getmilestone_dates_pss()
    for point in points:
        if '::' not in point:
            continue
        milestone = point.split('::')
        perc = milestone[0]
        descr = milestone[1]
        sheet['A' + str(i+1)] = i
        sheet['B' + str(i+1)] = descr
        sheet['C' + str(i+1)] = perc
        sheet['D' + str(i+1)] = 'PENDING'
        if i in dates:
            sheet['E' + str(i+1)] = dates[i]
            #sheet['E' + str(i+1)].number_format = 'M/D/YYYY'

        i += 1

    wb.save(project_path)
    return "Successfully retrieved payment milestone data for: " + project_path
    

def get_payment_milestones(project_path):
    data = [
        ['No.', 'Milestone Description', '%', 'Status', 'Forecast Date']
    ]
    if(os.path.exists(project_path) == False):
        return data
    wb = openpyxl.load_workbook(project_path)
    sheet_name = 'Payments'
    if sheet_name not in wb.sheetnames:
        return data
    filename = project_path
    # Load the Excel file
    df = pd.read_excel(filename, sheet_name=sheet_name,engine='openpyxl')
    for index, row in df.iterrows():
        num = str(row.iloc[0]) if pd.notna(row.iloc[0]) else ''
        descr = str(row.iloc[1]) if pd.notna(row.iloc[1]) else ''
        perc = str(row.iloc[2]) if pd.notna(row.iloc[2]) else ''
        status = str(row.iloc[3]) if pd.notna(row.iloc[3]) else ''
        forecast = row.iloc[4] if pd.notna(row.iloc[4]) else ''
        if isinstance(forecast, datetime):
            forecast = forecast.strftime("%m/%d/%Y")
        datum = [num, descr, perc, status, forecast]
        data.append(datum)

    return data


def write_delivery_data(project_path, po_data):
    if(os.path.exists(project_path) == False):
        return "File not found"
    wb = openpyxl.load_workbook(project_path)
    sheet_name = 'Delivery'
    if sheet_name not in wb.sheetnames:
        wb.create_sheet(sheet_name)
    sheet = wb[sheet_name]
    date = po_data[1]
    prod_name = po_data[0]
    if date.endswith('TERMINATE'):
        date = date[:-len('TERMINATE')]
    if prod_name.endswith('TERMINATE'):
        prod_name = prod_name[:-len('TERMINATE')]
    sheet['A1'] = 'No.'
    sheet['B1'] = 'Items'
    sheet['C1'] = 'Qty'
    sheet['D1'] = 'Unit'
    sheet['E1'] = 'Shipping Details'
    sheet['F1'] = 'Forecasted RTS'
    sheet['G1'] = 'Actual RTS'
    i = 1
    dates = getmilestone_dates(project_path)
    
    #item = point.split('::')
    #date = item[0]
    #descr = item[1]
    sheet['A' + str(i+1)] = i
    sheet['B' + str(i+1)] = prod_name
    sheet['F' + str(i+1)] = date
    i += 1

    wb.save(project_path)
    return "Successfully retrieved Scheduled Delivery data for: " + project_path


def get_delivery_data(project_path):
    data = [
        ['No.', 'Items', 'Qty', 'Unit', 'Shipping Details', 'Forecasted RTS', 'Actual RTS']
    ]
    if(os.path.exists(project_path) == False):
        return data
    wb = openpyxl.load_workbook(project_path)
    sheet_name = 'Delivery'
    if sheet_name not in wb.sheetnames:
        return data
    filename = project_path
    # Load the Excel file
    df = pd.read_excel(filename, sheet_name=sheet_name,engine='openpyxl')
    for index, row in df.iterrows():
        num = str(row.iloc[0]) if pd.notna(row.iloc[0]) else ''
        items = str(row.iloc[1]) if pd.notna(row.iloc[1]) else ''
        qty = str(row.iloc[2]) if pd.notna(row.iloc[2]) else ''
        unit = str(row.iloc[3]) if pd.notna(row.iloc[3]) else ''
        ship_dets = str(row.iloc[4]) if pd.notna(row.iloc[4]) else ''
        forecast_rts = row.iloc[5] if pd.notna(row.iloc[5]) else ''
        actual_rts = row.iloc[6] if pd.notna(row.iloc[6]) else ''
        if isinstance(forecast_rts, datetime):
            forecast_rts = forecast_rts.strftime("%m/%d/%Y")
        if isinstance(actual_rts, datetime):
            actual_rts = actual_rts.strftime("%m/%d/%Y")
        datum = [num, items, qty, unit, ship_dets, forecast_rts, actual_rts]
        data.append(datum)

    return data
#Usage
#filename = 'Task_Dependencies.xlsx'
#task_list = read_excel_tasks(filename)
#print(task_list)

#filename = "propel_beta\\Sample.xlsx"
#task_list = get_payment_milestones(filename)
#print(task_list)