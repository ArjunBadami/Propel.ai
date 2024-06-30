import autogen
from typing_extensions import Annotated
from .cpm import cpmcalc
from .excel_parser import read_excel_tasks, write_critical_start_times, write_payment_milestones, write_delivery_data
import json
import pandas as pd
import openpyxl
from autogen.agentchat.contrib.retrieve_user_proxy_agent import RetrieveUserProxyAgent
import numpy as np
from chromadb.utils import embedding_functions
import random
from chromadb.utils import embedding_functions
import PyPDF2
import chromadb
from .report_generator import create_report
from .cpm import cpmcalc

#default_ef = embedding_functions.DefaultEmbeddingFunction()
openai_ef = embedding_functions.OpenAIEmbeddingFunction(
                api_key="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",
                #api_key: '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' 
                model_name="text-embedding-ada-002"
            )


config_list = [
  {
      #'model': 'gpt-3.5-turbo-16k',
      'model': 'gpt-4',
      #'api_key': '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' 
      'api_key': "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" 
  }
]

llm_config = {
  "timeout": 600,
  "seed": 42,
  "config_list": config_list,
  "temperature": 0
}


def termination_msg(x):
    return isinstance(x, dict) and "TERMINATE" == str(x.get("content", ""))[-9:].upper()

boss = autogen.UserProxyAgent(
    name="User",
    is_termination_msg=termination_msg,
    human_input_mode="NEVER",
    code_execution_config=False,  # we don't want to execute code in this case.
    default_auto_reply="TERMINATE",
    description="The user who ask questions and give tasks.",
)


boss_aid = RetrieveUserProxyAgent(
    name="User_AID",
    is_termination_msg=termination_msg,
    human_input_mode="NEVER",
    max_consecutive_auto_reply=3,
    retrieve_config={
        "task": "code",
        "docs_path": "./po7.pdf",
        "chunk_token_size": 1000,
        "model": config_list[0]["model"],
        "embedding_function": openai_ef,
        "client": chromadb.PersistentClient(path="./chromadb"),
        "collection_name": "groupchat",
        "get_or_create": True,
    },
    code_execution_config=False,  # we don't want to execute code in this case.
    #default_auto_reply="TERMINATE",
    description="Assistant who has extra content retrieval power for solving difficult problems.",
)


'''
coder = autogen.AssistantAgent(
    name="Data_Manager",
    is_termination_msg=termination_msg,
    system_message="You are a expert data manager, you need to convert a given data to the request format. Don't do any other task. If asked you need to write into a file. Before writing, make sure to convert the data to json format with a list of objects without changing any text. Reply `TERMINATE` in the end when everything is done.",
    llm_config={"config_list": config_list, "timeout": 60, "temperature": 0},
    description="Data manager who can convert any data to structured data",
)
3. write_data_to_excel(<file_name>) -  Suggest this operation for any prompts including writing to excel.
4. excel_reader(<file_name>) - Suggest this operation for any prompts including reading from an excel.
'''

coder = autogen.AssistantAgent(
    name="Data_Manager",
    is_termination_msg=termination_msg,
    system_message="""You are responsible to suggest the name of an operation and its argument to do based on the task given. You do no anything else and reply just with the operation name.
    The possible options are(in order of first match): 
    1. getMonthlyReport(<file_name>) - Suggest this operation for any prompts to create a monthly report
    2. getcpm(<file_name>) - Suggest this operation for any prompts including Critical Path.
    3. getPaymentMilestones(<file_name>) - Suggest this operation for any prompts including Payment Milestone UNLESS it begins with `Retrieve Content:`
    4. retrieve_respond("<full_prompt_message>") - Suggest this operation for all prompts beginning with the phrase `From Purchase Order:`
    """,
    llm_config={"config_list": config_list, "timeout": 60, "temperature": 0},
    description="Data manager who can convert any data to structured data",
)


pm = autogen.AssistantAgent(
    name="Product_Manager",
    is_termination_msg=termination_msg,
    system_message="You are a product manager. Reply `TERMINATE` in the end when everything is done.",
    llm_config={"config_list": config_list, "timeout": 60, "temperature": 0},
    description="Product Manager who can design and plan the project.",
    #default_auto_reply="TERMINATE",
)



def generate_random_numbers(n):
  # Generate n random numbers from a Dirichlet distribution
  # The parameters are all set to 1, which is equivalent to a uniform distribution
  random_numbers = np.random.dirichlet(np.ones(n))
  return random_numbers

def write_data_to_excel(excel_file_path, data):
  try:
      json_data = json.loads(data)
      df = pd.DataFrame(json_data)
      #print(json_data)
      wb = openpyxl.Workbook()
      sheet = wb.active
      row = 2
      column = 1
      task_id = 10

      sheet.cell(row=1, column=1, value='Task Number')
      sheet.cell(row=1, column=2, value='Name')
      sheet.cell(row=1, column=3, value='Start Date')
      sheet.cell(row=1, column=4, value='Depends On')
      sheet.cell(row=1, column=5, value='Duration (weeks)')

    #index = 0
    #for ms in df['Milestone']:
      for index in range(len(json_data)):
          ms = json_data[index]['Milestone']
          sheet.cell(row=row, column=1, value=str(task_id))
          if str(ms).strip() == '':
            ms = 'Task ' + str(task_id)
          sheet.cell(row=row, column=2, value=str(ms))
          sheet.cell(row=row, column=3, value='')
          sheet.cell(row=row, column=4, value='')
          ms_dur = random.randint(1, 50)
          sheet.cell(row=row, column=5, value=ms_dur)
          row += 1
          ms_id = task_id
          task_id += 10
          #conds = df['Conditions'][index][1:-1].split('\', \'')
          conds = []
          if 'Conditions' in json_data[index]:
            conds = json_data[index]['Conditions']
          elif 'Condition' in json_data[index]:
            conds = json_data[index]['Condition']

          if isinstance(conds, str):
            conds = [conds]
        
          n = len(conds)
          random_numbers = generate_random_numbers(n)
          r_ind = 0
          for cond in conds:
              sheet.cell(row=row, column=1, value=str(task_id))
              if str(cond).strip() == '':
                cond = 'Task ' + str(task_id)
              sheet.cell(row=row, column=2, value=cond)
              sheet.cell(row=row, column=3, value='')
              sheet.cell(row=row, column=4, value=str(ms_id))
              sheet.cell(row=row, column=5, value=str(int(random_numbers[r_ind] * 100)))
              row += 1
              task_id += 10
              r_ind += 1

          #index += 1

      wb.save(excel_file_path)
  except Exception as error:
      return f"Sorry there was an error writing the file. Here is the data: \n {error}"
  return f"Written data to file {excel_file_path}"


#@boss.register_for_execution()
#@coder.register_for_llm(description="write excel")
def excel_writer(
    file: Annotated[str, "File path"],
    data: Annotated[str, "data"],
) -> str:
    return write_data_to_excel(file, data)


#@boss.register_for_execution()
#@coder.register_for_llm(description="Takes the name of the excel file as input. As output, it provides the list of task including task IDs, task names, task durations, and task dependencies as a formatted python list.")
def excel_reader(
    file: Annotated[str, "File path"],
) -> list:
    # df = pd.read_excel(file)
    # json_data = df.to_json(orient='records')
    # return json_data

    # Load the Excel file
    df = pd.read_excel(file, engine='openpyxl')

    # Initialize an empty list to store task data
    tasks = []

    # Iterate over each row in the DataFrame
    for index, row in df.iterrows():
        # Assuming the structure: Task ID | Task Name | Task Duration | Dependencies
        task_id = row.iloc[0]  # Convert Task ID to integer
        task_name = str(row.iloc[1])  # Task Name as string
        task_duration = row.iloc[4]  # Task Duration as integer
        dependencies = [int(float(v)) for v in str(row.iloc[3]).split(',')] if (pd.notna(row.iloc[3])
                                                                         and str(row.iloc[3]).strip() != '') else []

        # Append the task as a list to the tasks list
        tasks.append([task_id, task_name, task_duration, dependencies])

    return tasks


'''
@boss.register_for_execution()
@pm.register_for_llm(description="Critical Path of list of tasks.")
def getcpm(
      l: Annotated[list, "List of tasks. Each task is also a list where the first element is task ID, second element"
                         " is task name, 3rd element is task duration, and the 4th element is a list of task IDs that"
                         " are dependencies of this task."]
) -> str:
  return cpmcalc(l)
'''

def getcpm(filepath):
    project_tasks = read_excel_tasks(filepath)
    critical_path = cpmcalc(project_tasks)
    return write_critical_start_times(filepath, critical_path)


def extract_text_from_pdf(pdf_file_path):
  text = ""
  with open(pdf_file_path, 'rb') as file:
      reader = PyPDF2.PdfReader(file)
      num_pages = len(reader.pages)
      for page_num in range(num_pages):
          page = reader.pages[page_num]
          text += page.extract_text()
  
  return text


#@boss.register_for_execution()
#@coder.register_for_llm(description="Reads pdf file.")
def pdf_parser(
  pdf_file_path: Annotated[str, "File path of pdf to be read."]) -> str:
  return extract_text_from_pdf(pdf_file_path)

'''
def retrieve_content(
  message: Annotated[
      str,
      "Refined message which keeps the original meaning and can be used to retrieve content for code generation and question answering.",
  ],
  n_results: Annotated[int, "number of results"] = 3,
) -> str:
  boss_aid.n_results = n_results  # Set the number of results to be retrieved.
  # Check if we need to update the context.
  update_context_case1, update_context_case2 = boss_aid._check_update_context(message)
  if (update_context_case1 or update_context_case2) and boss_aid.update_context:
      boss_aid.problem = message if not hasattr(boss_aid, "problem") else boss_aid.problem
      _, ret_msg = boss_aid._generate_retrieve_user_reply(message)
  else:
      _context = {"problem": message, "n_results": n_results}
      ret_msg = boss_aid.message_generator(boss_aid, None, _context)
  return ret_msg if ret_msg else message

for caller in [coder]:
  d_retrieve_content = caller.register_for_llm(
      description="retrieve content for code generation and question answering.", api_style="tool"
  )(retrieve_content)

for executor in [boss]:
  executor.register_for_execution()(d_retrieve_content)

groupchat = autogen.GroupChat(
    agents=[boss, pm, coder], messages=[]
)
'''

def _reset_agents():
    boss.reset()
    boss_aid.reset()
    coder.reset()
    pm.reset()

'''
def rag_chat():
    _reset_agents()
    groupchat = autogen.GroupChat(
        agents=[boss_aid, pm, coder], messages=[]
    )
    manager = autogen.GroupChatManager(
        groupchat=groupchat, llm_config={"config_list": config_list, "timeout": 60, "temperature": 0}
    )

    # Start chatting with boss_aid as this is the user proxy agent.
    messages = boss_aid.initiate_chat(
        manager,
        message=boss_aid.message_generator,
        problem=PROBLEM,
        n_results=3,
        clear_history=False
    )

    return messages
'''

def call_chat(PROBLEM):
    _reset_agents()

    groupchat = autogen.GroupChat(
        agents=[boss, coder],
        messages=[],
        max_round=12,
        #speaker_selection_method="round_robin",
        allow_repeat_speaker=False,
    )

    manager = autogen.GroupChatManager(
        groupchat=groupchat, llm_config={"config_list": config_list, "timeout": 60, "temperature": 0}
    )

    # Start chatting with the boss as this is the user proxy agent.
    messages = boss.initiate_chat(
        manager,
        message=PROBLEM,
        clear_history=False
    )
    return messages



def call_rag_chat_included(PROBLEM):
    _reset_agents()

    # In this case, we will have multiple user proxy agents and we don't initiate the chat
    # with RAG user proxy agent.
    # In order to use RAG user proxy agent, we need to wrap RAG agents in a function and call
    # it from other agents.
    def retrieve_content(
        message: Annotated[
            str,
            "Refined message which keeps the original meaning and can be used to retrieve content for code generation and question answering.",
        ],
        n_results: Annotated[int, "number of results"] = 3,
    ) -> str:
        boss_aid.n_results = n_results  # Set the number of results to be retrieved.
        # Check if we need to update the context.
        update_context_case1, update_context_case2 = boss_aid._check_update_context(message)
        if (update_context_case1 or update_context_case2) and boss_aid.update_context:
            boss_aid.problem = message if not hasattr(boss_aid, "problem") else boss_aid.problem
            _, ret_msg = boss_aid._generate_retrieve_user_reply(message)
        else:
            _context = {"problem": message, "n_results": n_results}
            ret_msg = boss_aid.message_generator(boss_aid, None, _context)
        return ret_msg if ret_msg else message

    boss_aid.human_input_mode = "NEVER"  # Disable human input for boss_aid since it only retrieves content.

    for caller in [pm]:
        d_retrieve_content = caller.register_for_llm(
            description="retrieve content for question answering.", api_style="tool"
        )(retrieve_content)

    for executor in [boss]:
        executor.register_for_execution()(d_retrieve_content)

    groupchat = autogen.GroupChat(
        agents=[boss, pm],
        messages=[],
        max_round=12,
        #speaker_selection_method="round_robin",
        allow_repeat_speaker=False,
    )

    manager = autogen.GroupChatManager(
        groupchat=groupchat, llm_config={"config_list": config_list, "timeout": 60, "temperature": 0}
    )

    # Start chatting with the boss as this is the user proxy agent.
    messages = boss.initiate_chat(
        manager,
        message=PROBLEM,
        clear_history=False
    )
    return messages


def get_delivery_and_milestone_data():
    prompts = ["From \"po.pdf\", extract milestones for payment and invoicing. Paraphrase into numbered bullet points, each starting with `##`. Stick to only the points that describe the percentage of the total payment, keep it very brief, 1 phrase each. Begin each point with the percentage amount, followed by `::`"]
    results = []
    for prompt in prompts:
        message = call_rag_chat_included(prompt)
        last_non_empty_content = None
        for item in reversed(message.chat_history):
            if item['content'].strip() != '' and item['content'].strip() != 'TERMINATE' and item['content'].strip() != 'Reply `TERMINATE` if the task is done.':
                last_non_empty_content = item['content']
            #else:
                break
        
        results.append(str(last_non_empty_content))

    return results



def retrieve_respond(query):
    message = call_rag_chat_included(query)
    last_non_empty_content = None
    for item in reversed(message.chat_history):
        if item['content'].strip() != '' and item['content'].strip() != 'TERMINATE' and item['content'].strip() != 'Reply `TERMINATE` if the task is done.':
            last_non_empty_content = item['content']
        #else:
            break
        

    return last_non_empty_content


def getPaymentMilestones(filepath):
    po_data = get_delivery_and_milestone_data()
    return write_payment_milestones(filepath, po_data)


def get_scheduled_delivery_data():
    prompts = ["From \"po.pdf\", extract the product name from the document title. Return nothing else in your response apart from the name of the product(s)",
               "From the purchase order, what is the due date for the delivery of goods? Only return the due date, and no other words."]
    results = []
    for prompt in prompts:
        message = call_rag_chat_included(prompt)
        last_non_empty_content = None
        for item in reversed(message.chat_history):
            if item['content'].strip() != '' and item['content'].strip() != 'TERMINATE' and item['content'].strip() != 'Reply `TERMINATE` if the task is done.':
                last_non_empty_content = item['content']
            #else:
                break
        
        results.append(str(last_non_empty_content))

    return results


def getDeliveryData(filepath):
    po_data = get_scheduled_delivery_data()
    return write_delivery_data(filepath, po_data)


def getMonthlyReport(filepath):
   #po_data = get_delivery_and_milestone_data()
   res = getDeliveryData(filepath=filepath)
   po_data = None
   return create_report(filepath, po_data)